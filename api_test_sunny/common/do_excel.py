# _*_coding:utf-8 _*_
# @Time     :2018/12/30 21:44
# @Author   :sunny
# @Email    :602992114@qq.com

import json
import os
from openpyxl import load_workbook
from api_test_sunny.common.request import Requests
from api_test_sunny.common import constant

class Case:

    # 测试封装类
    def __init__(self):
        self.case_id=None
        self.case_description = None
        self.method = None
        self.url = None
        self.param = None
        self.expected = None
        self.actual = None
        self.result = None

class DoExcel:

    # Excel操作封装类
    def __init__(self,filename):
        try:
            self.filename = filename
            self.workbook = load_workbook(filename=filename)
        except  FileNotFoundError as e:
            print("报错的信息是：",e)
            raise e

    def get_cases(self,sheet_name):
        sheet=self.workbook[sheet_name]
        max_row=sheet.max_row # 获取所在sheet页的最大行
        cases=[]
        for i in range(2,max_row+1):
            case=Case() # 实例化一个case对象，用来存放测试数据
            case.case_id = sheet.cell(row=i,column=1).value
            case.case_description = sheet.cell(row=i,column=2).value
            case.method = sheet.cell(row=i,column=3).value
            case.url = sheet.cell(row=i,column=4).value
            case.param = sheet.cell(row=i,column=5).value
            case.expected = sheet.cell(row=i,column=6).value
            cases.append(case)
        return  cases

    def get_sheet_names(self): # 获取到workbook里面所有sheet名称的列表
        return self.workbook.sheetnames

    # 根据sheet name定位到sheet，然后根据case id 定位到行，取到当前行里面的actual这个单元格，
    # 然后给它赋值，最后保存当前workbook
    def write_text_by_case_id(self,sheet_name,case_id,actual,result):
        sheet=self.workbook[sheet_name]
        max_row=sheet.max_row
        for i in range(2,max_row+1):
            case_id_i=sheet.cell(i,1).value # 获取第 i 行第一列，也就是case_id这一列
            if case_id_i==case_id: # 判断取到Excel里面渠道的当前行的case_id 是否等于传进来的case_id
                sheet.cell(i,7).value=actual # 写入传进来的actual到当前的actual列的单元格
                sheet.cell(i,8).value=result # 写入传进来的actual到当前的result列的单元格
                self.workbook.save(self.filename)
                break

if __name__ == '__main__':
    print("coming*************")
    filename=constant.test_data_dir
    do_excel=DoExcel(filename)
    sheet_names=do_excel.get_sheet_names()
    print("sheet名称列表：",sheet_names)
    test_list=["register","login"] # 定义一个执行测试用例的列表
    for sheet_name in sheet_names:
        if sheet_name in test_list: # 如果当前的sheet name 不在可执行的test_list里面，就不执行这个sheet页里面的用例
            cases=do_excel.get_cases(sheet_name)
            print("case用例个数：",len(cases))
            for case  in cases:
                print("case信息:",case.__dict__)
                data=json.loads(case.param)# case.param 是一个字符串，用eval变成字典
                res=Requests(method=case.method,url=case.url,param=data,cookies=None,headers=None)
                print("status_code:",res.get_status_code())
                res_dict=res.get_json()
                # 通过json.dumps函数将字典转换成格式化后的字符串
                res_text=json.dumps(res_dict,ensure_ascii=False,indent=4)
                print("response:",res_text)
                if case.expected==res.get_text():
                    print("pass")
                    do_excel.write_text_by_case_id(sheet_name=sheet_name,case_id=case.case_id,actual=res.get_text(),
                                                   result="pass")
                else:
                    print("fail")
                    do_excel.write_text_by_case_id(sheet_name=sheet_name, case_id=case.case_id, actual=res.get_text(),
                                                   result="fail")

    # eval 函数不能自动识别null并转换成None
    # json 可以将字符串和dict的转换，推荐使用json